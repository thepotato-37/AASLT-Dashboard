from __future__ import annotations

import json
import re
from hashlib import sha1
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
LRTC_SOURCE_FILE = next(
    path
    for path in [
        Path("/Users/anantsabata/Downloads/Total AASLT Cadre LRTC (2).xlsx"),
        Path("/Users/anantsabata/Downloads/Total AASLT Cadre LRTC (1).xlsx"),
        Path("/Users/anantsabata/Downloads/Total AASLT Cadre LRTC.xlsx"),
    ]
    if path.exists()
)

SOURCE_FILES = [
    LRTC_SOURCE_FILE,
    Path("/Users/anantsabata/Downloads/1. CST 26 Mess Matrix (Live Document).xlsx"),
]

YEAR = 2026
MONTHS = {
    "JAN": 1,
    "JANUARY": 1,
    "FEB": 2,
    "FEBRUARY": 2,
    "MAR": 3,
    "MARCH": 3,
    "APR": 4,
    "APRIL": 4,
    "MAY": 5,
    "JUN": 6,
    "JUNE": 6,
    "JUL": 7,
    "JULY": 7,
    "AUG": 8,
    "AUGUST": 8,
    "SEP": 9,
    "SEPT": 9,
    "SEPTEMBER": 9,
    "OCT": 10,
    "OCTOBER": 10,
    "NOV": 11,
    "NOVEMBER": 11,
    "DEC": 12,
    "DECEMBER": 12,
}

MEAL_LABELS = {
    "BREAKFAST": "Breakfast",
    "LUNCH": "Lunch",
    "DINNER": "Dinner",
}
MEAL_ORDER = ("BREAKFAST", "LUNCH", "DINNER")
WP_CLASS_KEYS = {
    "WP700": "Air Assault 1",
    "WP701": "Air Assault 2",
    "WP702": "Air Assault 3",
    "WP703": "Air Assault 4",
    "WP704": "Air Assault 5",
}


def cell_to_json(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return value


def compact_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H%M")
    return re.sub(r"\s+", " ", str(value)).strip()


def parse_day_label(text: str, fallback_month: int | None = 6) -> str | None:
    upper = text.upper().replace(".", "")
    month_pattern = r"(JANUARY|JAN|FEBRUARY|FEB|MARCH|MAR|APRIL|APR|MAY|JUNE|JUN|JULY|JUL|AUGUST|AUG|SEPTEMBER|SEPT|SEP|OCTOBER|OCT|NOVEMBER|NOV|DECEMBER|DEC)"
    for month_name, day_text in re.findall(rf"\b{month_pattern}\s+(\d{{1,2}})\b", upper):
        month = MONTHS.get(month_name)
        day = int(day_text)
        if month:
            return date(YEAR, month, day).isoformat()
    for day_text, month_name in re.findall(rf"\b(\d{{1,2}})(?:ST|ND|RD|TH)?[-\s]+{month_pattern}\b", upper):
        month = MONTHS.get(month_name)
        day = int(day_text)
        if month:
            return date(YEAR, month, day).isoformat()
    match = re.search(r"\b(\d{1,2})(?:ST|ND|RD|TH)\b", upper)
    if match and fallback_month:
        day = int(match.group(1))
        if 1 <= day <= 31:
            return date(YEAR, fallback_month, day).isoformat()
    return None


def ordinal_day_from_text(text: str) -> int | None:
    match = re.search(r"\b(\d{1,2})(?:ST|ND|RD|TH)\b", compact_text(text).upper())
    if not match:
        return None
    day = int(match.group(1))
    return day if 1 <= day <= 31 else None


def date_to_iso(value: Any, sheet_title: str = "") -> str | None:
    if isinstance(value, datetime):
        day = value.date()
    elif isinstance(value, date):
        day = value
    else:
        return parse_day_label(compact_text(value), 6)
    title_year = re.search(r"\b(20\d{2})\b", sheet_title)
    if title_year:
        day = date(int(title_year.group(1)), day.month, day.day)
    return day.isoformat()


def parse_time_token(text: str) -> tuple[str | None, str | None]:
    upper = text.upper().replace("–", "-").replace("—", "-")
    match = re.search(r"\b([0-2]?\d[0-5]\d)\s*-\s*([0-2]?\d[0-5]\d|UTC)\b", upper)
    if match:
        return normalize_time(match.group(1)), normalize_time(match.group(2))
    match = re.search(r"\b([0-2]?\d[0-5]\d)\b", upper)
    if match:
        return normalize_time(match.group(1)), None
    return None, None


def normalize_time(token: str) -> str | None:
    if token.upper() == "UTC":
        return None
    digits = token.zfill(4)
    hour = int(digits[:2])
    minute = int(digits[2:])
    if hour > 23 or minute > 59:
        return None
    return f"{hour:02d}:{minute:02d}"


def infer_category(text: str, header: str = "") -> str:
    upper = f"{header} {text}".upper()
    if any(term in upper for term in ["MESS", "MEAL", "BREAKFAST", "LUNCH", "DINNER", "MERMITE"]):
        return "Meals"
    if "MEDIC" in upper or "FLA" in upper or "KACH" in upper:
        return "Medical"
    if "S4" in upper or "WATER BUFFALO" in upper or "DRAW" in upper or "SUPPLY" in upper:
        return "Logistics"
    if "CADRE" in upper or "CSM" in upper or "S1" in upper:
        return "Cadre"
    if "AA1" in upper or "AA2" in upper or "AA3" in upper or "TRAINEE" in upper:
        return "Training"
    if "RUCK" in upper or "TEST" in upper or "RAPPEL" in upper or "SLING" in upper:
        return "Training"
    if "CLASS" in upper or "BRIEF" in upper or "ORIENTATION" in upper:
        return "Instruction"
    return "Operations"


def is_non_ruck_fla_text(text: str) -> bool:
    upper = compact_text(text).upper()
    return "FLA" in upper and "RUCK" not in upper


def infer_location(text: str) -> str:
    candidates = [
        "Bartlett",
        "Davis Shelf",
        "South Dock",
        "Washington Hall",
        "Camp Buckner",
        "Trophy Point",
        "KACH",
        "Grad Pad",
        "Ruck AXP 1",
        "Ruck AXP 2",
        "4x36",
        "The Plain",
    ]
    lower = text.lower()
    for candidate in candidates:
        if candidate.lower() in lower:
            return candidate
    return ""


def infer_class_key(*parts: str) -> str | None:
    text = " ".join(compact_text(part) for part in parts if part)
    upper = text.upper()
    match = re.search(r"\bAIR\s*ASSAULT\s*([1-9]\d*)\b", upper)
    if match:
        return f"Air Assault {int(match.group(1))}"
    match = re.search(r"\bAASLT\s*([1-9]\d*)\b", upper)
    if match:
        return f"Air Assault {int(match.group(1))}"
    match = re.search(r"\bAA\s*([1-9]\d*)\b", upper)
    if match:
        return f"Air Assault {int(match.group(1))}"
    for wp_code, class_key in WP_CLASS_KEYS.items():
        if re.search(rf"\b{wp_code}\b", upper):
            return class_key
    return None


def day_number_from_label(text: str) -> int | None:
    upper = compact_text(text).upper()
    if upper == "ZERO":
        return 0
    try:
        return int(upper)
    except ValueError:
        return None


def day_numbers_from_header(text: str) -> list[int]:
    numbers: list[int] = []
    for token in re.findall(r"\bDAY\s+(-?\d+|ZERO)\b", compact_text(text).upper()):
        number = day_number_from_label(token)
        if number is not None:
            numbers.append(number)
    return numbers


def build_class_by_date_day(events: list[dict[str, Any]]) -> dict[tuple[str, int], str]:
    lookup: dict[tuple[str, int], str] = {}
    for event in events:
        event_date = event.get("date")
        if not event_date:
            continue
        text = " ".join(compact_text(event.get(key)) for key in ("title", "group", "notes"))
        match = re.search(r"\bDAY\s+(-?\d+|ZERO)\s+WP\d+\b", text, re.I)
        if not match:
            continue
        day_number = day_number_from_label(match.group(1))
        if day_number is None:
            continue
        class_key = event.get("classKey") or infer_class_key(text)
        if class_key:
            lookup[(event_date, day_number)] = class_key
    anchors = [(iso, day_number, class_key) for (iso, day_number), class_key in lookup.items()]
    for iso, anchor_day, class_key in anchors:
        anchor_date = date.fromisoformat(iso)
        for target_day in range(-3, 11):
            target_date = anchor_date + timedelta(days=target_day - anchor_day)
            if target_date.year != YEAR:
                continue
            lookup.setdefault((target_date.isoformat(), target_day), class_key)
    return lookup


def cell_font_color_key(cell: Any) -> str:
    color = cell.font.color
    if not color:
        return "default"
    if color.type == "rgb" and color.rgb:
        return color.rgb.upper()
    if color.type == "indexed" and color.indexed is not None:
        return f"indexed:{color.indexed}"
    if color.type == "theme" and color.theme is not None:
        return f"theme:{color.theme}"
    return color.type or "default"


def is_second_medical_class_color(color_key: str) -> bool:
    return color_key in {"FFFF0000", "00FF0000", "FF0000"}


def medical_owner_from_color(
    *,
    event_date: str,
    header_text: str,
    value_cells: list[Any],
    class_by_date_day: dict[tuple[str, int], str],
) -> tuple[str | None, int | None, str]:
    day_numbers = day_numbers_from_header(header_text)
    color_key = next((cell_font_color_key(cell) for cell in value_cells if is_second_medical_class_color(cell_font_color_key(cell))), cell_font_color_key(value_cells[0]))
    day_index = 1 if len(day_numbers) > 1 and is_second_medical_class_color(color_key) else 0
    medical_day = day_numbers[day_index] if day_numbers else None
    class_key = class_by_date_day.get((event_date, medical_day)) if medical_day is not None else None
    if not class_key and day_numbers:
        class_key = class_by_date_day.get((event_date, day_numbers[0]))
    return class_key, medical_day, color_key


def meal_labels(text: str) -> list[str]:
    upper = text.upper()
    return [MEAL_LABELS[label] for label in MEAL_ORDER if re.search(rf"\b{label}\b", upper)]


def normalize_meal_title(text: str, source_kind: str) -> str:
    if source_kind == "mess":
        return text
    labels = meal_labels(text)
    if not labels:
        return text

    upper = text.upper()
    if re.search(r"\bMRE\s+BREAKFAST\s*\+\s*LUNCH\b", upper):
        return "Breakfast/Lunch MRE"
    if re.search(r"\bBREAKFAST\s*/\s*LUNCH\s+MRE\b", upper):
        return "Breakfast/Lunch MRE"

    if len(labels) == 1 and re.fullmatch(
        rf"\s*{labels[0].upper()}\s+[0-2]?\d[0-5]\d\s*-\s*(?:[0-2]?\d[0-5]\d|UTC)\s*",
        upper,
    ):
        return labels[0]

    suffix = ""
    if re.search(r"\bMRE\b", upper):
        suffix = " MRE"
    elif "FORMATION" in upper:
        suffix = " Formation"

    return "/".join(labels) + suffix


def normalize_ruck_medical_requirement(requirement: str, location: str, time_value: str) -> str:
    text = compact_text(requirement)
    context = f"{location} {time_value} {text}".upper()
    if "RUCK" not in context or "FLA" not in context:
        return text
    normalized = re.sub(r"\b(?:X\s*)?\d+\s*(?:X\s*)?FLA\b", "3 FLA", text, flags=re.I)
    if normalized == text:
        return "3 FLA W/ CREW" if "CREW" in context else "3 FLA"
    return normalized


def class_context_for_cell(ws: Any, row: int, col: int) -> tuple[str | None, str]:
    marker_row: int | None = None
    class_key: str | None = None
    for scan_row in range(row, max(1, row - 24) - 1, -1):
        text = compact_text(ws.cell(scan_row, col).value)
        class_key = infer_class_key(text)
        if class_key:
            marker_row = scan_row
            break
        if parse_day_label(text, 6) and scan_row != row:
            break

    location = ""
    if marker_row:
        for scan_row in range(marker_row, min(row, marker_row + 3) + 1):
            location = infer_location(compact_text(ws.cell(scan_row, col).value))
            if location:
                break
    return class_key, location


def make_event(
    *,
    events: list[dict[str, Any]],
    source_file: str,
    sheet: str,
    row: int,
    col: int,
    event_date: str | None,
    title: str,
    header: str = "",
    assigned: str = "",
    source_kind: str = "schedule",
    class_key: str | None = None,
    location: str = "",
    source_extra: dict[str, Any] | None = None,
) -> None:
    text = compact_text(title)
    if not event_date or not text:
        return
    if len(text) <= 1:
        return
    if re.fullmatch(r"\d{4}", text):
        return
    start, end = parse_time_token(text)
    notes = "" if source_kind == "class-calendar" else compact_text(assigned)
    combined = " ".join(part for part in [header, text, notes] if part)
    category = infer_category(combined, header)
    if category == "Meals":
        text = normalize_meal_title(text, source_kind)
    event_location = location or infer_location(combined)
    event_class_key = class_key or infer_class_key(combined, source_file, sheet)
    event_id = f"evt-{len(events) + 1:05d}"
    source = {
        "file": source_file,
        "sheet": sheet,
        "cell": f"{get_column_letter(col)}{row}",
        "row": row,
        "col": col,
    }
    if source_extra:
        source.update(source_extra)
    event_record = {
        "id": event_id,
        "date": event_date,
        "start": start,
        "end": end,
        "title": text,
        "category": category,
        "group": compact_text(header),
        "location": event_location,
        "people": [],
        "notes": notes,
        "classKey": event_class_key,
        "sourceKind": source_kind,
        "source": source,
    }
    if source_extra:
        for key in ("lane", "role", "dateHeader", "fontColor", "dayNumber"):
            if source_extra.get(key) is not None and source_extra.get(key) != "":
                event_record[key] = source_extra[key]
    events.append(event_record)


def lrtc_column_role(header: str) -> str:
    upper = compact_text(header).upper()
    if "CADRE" in upper:
        return "CADRE"
    if "ADMIN" in upper:
        return "ADMIN"
    if "MEDIC" in upper:
        return "MEDICS"
    if "TRAINEE" in upper:
        return "TRAINEES"
    if "HOURLY" in upper:
        return "HOURLY"
    return ""


def class_key_from_aa_number(number: str | int) -> str:
    return f"Air Assault {int(number)}"


def class_key_from_lrtc_header(header: str) -> str | None:
    match = re.search(r"\bAA\s*([1-9]\d*)\b", compact_text(header), re.I)
    return class_key_from_aa_number(match.group(1)) if match else None


def split_lrtc_entries(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    pieces = [compact_text(piece) for piece in text.split("\n")]
    pieces = [piece for piece in pieces if piece]
    if len(pieces) <= 1:
        return pieces
    timed_pieces = [piece for piece in pieces if parse_time_token(piece)[0]]
    # Support cells often wrap one requirement across several lines, e.g.
    # "0530-0845 / 1x FLA Bartlett / 1x FLA Davis Shelf". Keep those intact.
    if len(timed_pieces) < 2:
        return [compact_text(text)]
    return pieces


def cell_color_key(cell: Any) -> str:
    color = cell.font.color
    if not color:
        return ""
    if color.type == "rgb" and color.rgb:
        return color.rgb.upper()
    if color.type == "indexed" and color.indexed is not None:
        return f"indexed:{color.indexed}"
    if color.type == "theme" and color.theme is not None:
        return f"theme:{color.theme}"
    return color.type or ""


def is_red_lrtc_color(color: str) -> bool:
    return color in {"FFFF0000", "00FF0000", "FF0000"}


def is_blue_lrtc_color(color: str) -> bool:
    return color in {"FF4472C4", "FF0070C0", "000070C0", "004472C4"}


def is_green_lrtc_color(color: str) -> bool:
    return color in {"FF00B050", "0000B050", "FF70AD47"}


def day_track_map_for_block(date_header: str, block_headers: list[str]) -> dict[int, str]:
    trainee_tracks = []
    for header in block_headers:
        class_key = class_key_from_lrtc_header(header)
        if class_key and class_key not in trainee_tracks:
            trainee_tracks.append(class_key)
    day_numbers = day_numbers_from_header(date_header)
    if not trainee_tracks:
        return {}
    return {day_number: trainee_tracks[index] for index, day_number in enumerate(day_numbers[: len(trainee_tracks)])}


def infer_lrtc_track(
    *,
    text: str,
    header: str,
    date_header: str,
    block_headers: list[str],
    color: str,
) -> str | None:
    explicit = infer_class_key(text, header) or class_key_from_lrtc_header(header)
    if explicit:
        return explicit

    day_to_track = day_track_map_for_block(date_header, block_headers)
    if len(day_to_track) == 1:
        return next(iter(day_to_track.values()))

    # In the current LRTC convention, red/blue/green font colors distinguish
    # track-specific support cells inside shared admin/medic columns.
    ordered_tracks = [track for _, track in sorted(day_to_track.items())]
    if is_red_lrtc_color(color) and ordered_tracks:
        return ordered_tracks[0]
    if (is_blue_lrtc_color(color) or is_green_lrtc_color(color)) and len(ordered_tracks) > 1:
        return ordered_tracks[1]
    return None


def lrtc_date_blocks(ws: Any) -> list[dict[str, Any]]:
    date_by_col: dict[int, str] = {}
    date_header_by_col: dict[int, str] = {}
    for row_idx in range(1, min(ws.max_row, 8) + 1):
        for col_idx in range(1, ws.max_column + 1):
            text = compact_text(ws.cell(row_idx, col_idx).value)
            day = parse_day_label(text, 6)
            if day:
                date_by_col[col_idx] = day
                date_header_by_col[col_idx] = text
    blocks = []
    sorted_cols = sorted(date_by_col)
    for idx, start_col in enumerate(sorted_cols):
        end_col = (sorted_cols[idx + 1] - 1) if idx + 1 < len(sorted_cols) else ws.max_column
        block_headers = [compact_text(ws.cell(2, col).value) for col in range(start_col, end_col + 1)]
        blocks.append(
            {
                "startCol": start_col,
                "endCol": end_col,
                "date": date_by_col[start_col],
                "header": date_header_by_col[start_col],
                "blockHeaders": block_headers,
            }
        )
    return blocks


def build_lrtc_class_by_date_day(wb: Any) -> dict[tuple[str, int], str]:
    lookup: dict[tuple[str, int], str] = {}
    for ws in wb.worksheets:
        for block in lrtc_date_blocks(ws):
            for day_number, class_key in day_track_map_for_block(block["header"], block["blockHeaders"]).items():
                lookup[(block["date"], day_number)] = class_key

    anchors = [(iso, day_number, class_key) for (iso, day_number), class_key in lookup.items()]
    for iso, anchor_day, class_key in anchors:
        anchor_date = date.fromisoformat(iso)
        for target_day in range(-3, 15):
            target_date = anchor_date + timedelta(days=target_day - anchor_day)
            if target_date.year != YEAR:
                continue
            lookup.setdefault((target_date.isoformat(), target_day), class_key)
    return lookup


def lrtc_lane_for_role(role: str, header: str, text: str) -> str:
    upper = f"{header} {text}".upper()
    if role == "TRAINEES":
        return "Trainees"
    if role == "CADRE":
        return "Cadre"
    if role == "ADMIN":
        return "Admin"
    if role == "MEDICS":
        return "Medics"
    if "S4" in upper:
        return "S4"
    return role.title() if role else "Timeline"


def required_people_from_text(text: str, role: str) -> int:
    matches = [int(match) for match in re.findall(r"\b(\d+)\s*x\s*Cadre\b", text, re.I)]
    if matches:
        return max(1, sum(matches))
    if role in {"CADRE", "ADMIN"}:
        generic = re.search(r"\b(\d+)\s*x\b", text, re.I)
        if generic:
            return max(1, int(generic.group(1)))
    return 1


def taskees_from_text(text: str) -> int:
    match = re.search(r"\b(\d+)\s*taskees?\b", text, re.I)
    return int(match.group(1)) if match else 0


def is_lrtc_task_cell(text: str, role: str) -> bool:
    upper = compact_text(text).upper()
    if role in {"CADRE", "ADMIN"}:
        return True
    if upper in {"B", "Z"}:
        return True
    task_terms = [
        "LOGSYNC",
        "WATER BUFFALO",
        "MRE DISTRO",
        "ENABLER",
        "FLA REQUEST",
        "RECEIVE DV",
        "RECIEVE DV",
        "TASKEE BRIEF",
        "CADRE IN-BRIEF",
        "MAJ L",
        "MRE PICKUP",
        "MRE ISSUE",
        "LOADOUT",
        "ADMIN SYNC",
        "STAFF SYNC",
    ]
    return any(term in upper for term in task_terms)


def normalize_lrtc_task_title(text: str, role: str) -> str:
    cleaned = compact_text(text)
    if cleaned.upper() in {"B", "Z"}:
        return ""
    cleaned = re.sub(r"^\s*[0-2]?\d[0-5]\d\s*-\s*(?:[0-2]?\d[0-5]\d|UTC)\s*", "", cleaned, flags=re.I).strip()
    cleaned = re.sub(r"\b([0-2]?\d[0-5]\d)\b\s*", "", cleaned, count=1).strip()
    if role == "CADRE" and re.fullmatch(r"\d+\s*x\s*Cadre(?:\s+\d+\s*x\s*Cadre)*", cleaned, re.I):
        return f"Cadre coverage - {cleaned}"
    if re.search(r"\bMRE\s+DISTRO\b", cleaned, re.I):
        return re.sub(r"\bMRE\s+DISTRO\b", "MRE distro", cleaned, flags=re.I)
    if re.search(r"\bFLA\s+REQUEST\b", cleaned, re.I):
        return re.sub(r"^FLAG:\s*", "", cleaned, flags=re.I)
    return cleaned


def make_tasking(
    *,
    taskings: list[dict[str, Any]],
    source_file: str,
    sheet: str,
    row: int,
    col: int,
    event_date: str | None,
    title: str,
    role: str,
    header: str = "",
    class_key: str | None = None,
    row_start: str | None = None,
    lane: str = "",
    required_people: int | None = None,
    taskees: int = 0,
) -> None:
    text = compact_text(title)
    if not event_date or not text or len(text) <= 1:
        return
    if re.fullmatch(r"\d{4}", text):
        return
    start, end = parse_time_token(text)
    start = start or row_start
    task_title = normalize_lrtc_task_title(text, role)
    if not task_title or len(task_title) <= 1:
        return
    source = {
        "file": source_file,
        "sheet": sheet,
        "cell": f"{get_column_letter(col)}{row}",
        "row": row,
        "col": col,
    }
    taskings.append(
        {
            "id": stable_id("tasking", source_file, sheet, source["cell"], task_title),
            "date": event_date,
            "start": start,
            "end": end,
            "title": task_title,
            "role": role or "TASK",
            "category": infer_category(f"{header} {text}", header),
            "track": class_key or "Shared",
            "location": infer_location(f"{header} {text}"),
            "lane": lane or lrtc_lane_for_role(role, header, text),
            "requiredPeople": required_people or required_people_from_text(text, role),
            "taskees": taskees_from_text(text) if taskees == 0 else taskees,
            "notes": f"LRTC {role.lower()} cell".strip(),
            "sourceKind": "lrtc-tasking",
            "source": source,
        }
    )


def support_lane_from_text(text: str, role: str) -> str:
    upper = compact_text(text).upper()
    if "WATER BUFFALO" in upper:
        return "Water Buffalo"
    if "FLA" in upper or role == "MEDICS":
        return "FLA Request"
    if "TRANSPO" in upper or "TRANSPORT" in upper:
        return "Transportation"
    if "DRAW" in upper or "LOGSYNC" in upper or "SUPPLY" in upper or "S4" in upper:
        return "Draw"
    return ""


def support_quantity_from_text(text: str) -> int:
    match = re.search(r"\b(\d+)\s*x\b", compact_text(text), re.I)
    return int(match.group(1)) if match else 1


def make_support_item(
    *,
    support_items: list[dict[str, Any]],
    source_file: str,
    sheet: str,
    row: int,
    col: int,
    event_date: str | None,
    title: str,
    role: str,
    class_key: str | None = None,
    row_start: str | None = None,
) -> None:
    text = compact_text(title)
    if not event_date or not text or text.upper() in {"B", "Z"}:
        return
    lane = support_lane_from_text(text, role)
    if not lane:
        return
    start, _ = parse_time_token(text)
    start = start or row_start or "00:00"
    source = {
        "file": source_file,
        "sheet": sheet,
        "cell": f"{get_column_letter(col)}{row}",
        "row": row,
        "col": col,
    }
    support_items.append(
        {
            "id": stable_id("support", source_file, sheet, source["cell"], text),
            "lane": lane,
            "item": normalize_lrtc_task_title(text, role) or text,
            "qty": support_quantity_from_text(text),
            "datetime": f"{event_date}T{start}",
            "status": "Planned",
            "owner": "S4/AS4" if re.search(r"S4|AS4", text, re.I) else "",
            "track": class_key or "Shared",
            "notes": "Seeded from LRTC support cell",
            "sourceKind": "lrtc-support",
            "source": source,
        }
    )


def extract_raw_workbook(path: Path) -> dict[str, Any]:
    values_wb = load_workbook(path, data_only=True)
    formulas_wb = load_workbook(path, data_only=False)
    sheets = []
    for values_ws in values_wb.worksheets:
        formula_ws = formulas_wb[values_ws.title]
        cells = []
        for row in values_ws.iter_rows():
            for cell in row:
                formula_value = formula_ws[cell.coordinate].value
                value = cell_to_json(cell.value)
                formula = formula_value if isinstance(formula_value, str) and formula_value.startswith("=") else None
                if value is not None or formula:
                    cells.append(
                        {
                            "r": cell.row,
                            "c": cell.column,
                            "a": cell.coordinate,
                            "v": value,
                            "f": formula,
                        }
                    )
        sheets.append(
            {
                "name": values_ws.title,
                "maxRow": values_ws.max_row,
                "maxColumn": values_ws.max_column,
                "cells": cells,
            }
        )
    return {"name": path.name, "sheets": sheets}


def parse_classes(path: Path, events: list[dict[str, Any]]) -> None:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    month_pattern = r"\b(JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER|JAN|FEB|MAR|APR|JUN|JUL|AUG|SEP|SEPT|OCT|NOV|DEC)\b"
    month_by_cell: dict[tuple[int, int], int] = {}
    for row_idx in range(1, ws.max_row + 1):
        anchors: list[tuple[int, int]] = []
        for col_idx in range(1, ws.max_column + 1):
            text = compact_text(ws.cell(row_idx, col_idx).value).upper()
            month_match = re.search(month_pattern, text)
            if month_match:
                month = MONTHS.get(month_match.group(1), 6)
                anchors.append((col_idx, month))
        for idx, (start_col, month) in enumerate(anchors):
            end_col = anchors[idx + 1][0] - 1 if idx + 1 < len(anchors) else ws.max_column
            for col_idx in range(start_col, end_col + 1):
                month_by_cell[(row_idx, col_idx)] = month

    for row_idx in range(1, ws.max_row + 1):
        row_values = [compact_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        for col_idx, text in enumerate(row_values, start=1):
            current_month = 6
            for scan_row in range(row_idx, max(1, row_idx - 8) - 1, -1):
                month = month_by_cell.get((scan_row, col_idx))
                if month:
                    current_month = month
                    break
            day = parse_day_label(text, current_month)
            if not day:
                continue
            # Calendar-style sheets usually place one day across two columns:
            # the left column is the event/activity and the right column carries names.
            for data_row in range(row_idx + 1, min(row_idx + 90, ws.max_row + 1)):
                below_joined = " ".join(compact_text(ws.cell(data_row, c).value) for c in range(1, min(ws.max_column, 12) + 1)).upper()
                if data_row > row_idx + 12 and ("WEST POINT CALENDAR" in below_joined or re.fullmatch(r"JUNE\s+\d{1,2}-\d{1,2}", below_joined)):
                    break
                activity = compact_text(ws.cell(data_row, col_idx).value)
                assigned = compact_text(ws.cell(data_row, col_idx + 1).value) if col_idx + 1 <= ws.max_column else ""
                if activity and not parse_day_label(activity, current_month):
                    class_key, location = class_context_for_cell(ws, data_row, col_idx)
                    make_event(
                        events=events,
                        source_file=path.name,
                        sheet=ws.title,
                        row=data_row,
                        col=col_idx,
                        event_date=day,
                        title=activity,
                        header="Class Schedule",
                        assigned=assigned,
                        source_kind="class-calendar",
                        class_key=class_key,
                        location=location,
                    )


def parse_lrtc(path: Path, events: list[dict[str, Any]], taskings: list[dict[str, Any]], support_items: list[dict[str, Any]]) -> None:
    wb = load_workbook(path, data_only=True)
    style_wb = load_workbook(path, data_only=False)
    class_by_date_day = build_lrtc_class_by_date_day(wb)
    for ws in wb.worksheets:
        style_ws = style_wb[ws.title]
        date_by_col: dict[int, str] = {}
        date_header_by_col: dict[int, str] = {}
        header_by_col: dict[int, str] = {}
        role_by_col: dict[int, str] = {}
        block_headers_by_col: dict[int, list[str]] = {}
        for block in lrtc_date_blocks(ws):
            for col in range(block["startCol"], block["endCol"] + 1):
                date_by_col[col] = block["date"]
                date_header_by_col[col] = block["header"]
                block_headers_by_col[col] = block["blockHeaders"]

        for col_idx in range(1, ws.max_column + 1):
            header_parts = []
            for row_idx in range(1, min(ws.max_row, 2) + 1):
                text = compact_text(ws.cell(row_idx, col_idx).value)
                if text and not parse_day_label(text, 6) and text.upper() not in {"WK 2", "WEEK 3", "WEEK 4"}:
                    header_parts.append(text)
            header_by_col[col_idx] = " ".join(header_parts)
            role_by_col[col_idx] = lrtc_column_role(compact_text(ws.cell(2, col_idx).value))

        for row_idx in range(3, ws.max_row + 1):
            hourly = compact_text(ws.cell(row_idx, 1).value)
            row_start = normalize_time(hourly) if re.fullmatch(r"\d{3,4}", hourly) else None
            for col_idx in range(2, ws.max_column + 1):
                entries = split_lrtc_entries(ws.cell(row_idx, col_idx).value)
                role = role_by_col.get(col_idx, "")
                if role == "HOURLY":
                    continue
                event_date = date_by_col.get(col_idx)
                header = header_by_col.get(col_idx, "")
                if not entries or not event_date:
                    continue
                color = cell_color_key(style_ws.cell(row_idx, col_idx))
                block_headers = block_headers_by_col.get(col_idx, [])
                for entry_index, text in enumerate(entries, start=1):
                    if not text or parse_day_label(text, 6):
                        continue
                    if re.fullmatch(r"[BZ]", text, re.I):
                        continue
                    class_key = infer_lrtc_track(
                        text=text,
                        header=header,
                        date_header=date_header_by_col.get(col_idx, ""),
                        block_headers=block_headers,
                        color=color,
                    )
                    if not class_key:
                        for day_number in day_numbers_from_header(date_header_by_col.get(col_idx, "")):
                            class_key = class_by_date_day.get((event_date, day_number))
                            if class_key:
                                break
                    day_number = None
                    for candidate_day, candidate_track in day_track_map_for_block(date_header_by_col.get(col_idx, ""), block_headers).items():
                        if class_key and candidate_track == class_key:
                            day_number = candidate_day
                            break
                    if day_number is None and class_key:
                        for (lookup_date, lookup_day), lookup_track in class_by_date_day.items():
                            if lookup_date == event_date and lookup_track == class_key:
                                day_number = lookup_day
                                break
                    lane = lrtc_lane_for_role(role, header, text)
                    title = text
                    if row_start and not parse_time_token(title)[0]:
                        title = f"{hourly} {text}"
                    source_extra = {
                        "lane": lane,
                        "dateHeader": date_header_by_col.get(col_idx, ""),
                        "dayNumber": day_number,
                        "role": role,
                        "entryIndex": entry_index,
                        "fontColor": color,
                    }
                    if is_non_ruck_fla_text(text):
                        continue
                    make_event(
                        events=events,
                        source_file=path.name,
                        sheet=ws.title,
                        row=row_idx,
                        col=col_idx,
                        event_date=event_date,
                        title=title,
                        header=lane,
                        assigned="",
                        source_kind="lrtc",
                        class_key=class_key,
                        source_extra=source_extra,
                    )
                    if role in {"CADRE", "ADMIN"} or (role != "TRAINEES" and is_lrtc_task_cell(text, role)):
                        make_tasking(
                            taskings=taskings,
                            source_file=path.name,
                            sheet=ws.title,
                            row=row_idx,
                            col=col_idx,
                            event_date=event_date,
                            title=title,
                            role=role,
                            header=header,
                            class_key=class_key,
                            row_start=row_start,
                            lane=lane,
                            required_people=required_people_from_text(text, role),
                            taskees=taskees_from_text(text),
                        )


def parse_medical(path: Path, events: list[dict[str, Any]]) -> None:
    wb = load_workbook(path, data_only=True)
    style_wb = load_workbook(path, data_only=False)
    ws = wb.active
    style_ws = style_wb[ws.title]
    class_by_date_day = build_class_by_date_day(events)
    header_cells: list[tuple[int, int, str]] = []
    header_candidates: list[tuple[int, int, int]] = []
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            day_number = ordinal_day_from_text(compact_text(ws.cell(row_idx, col_idx).value))
            if day_number:
                header_candidates.append((row_idx, col_idx, day_number))

    current_month = 6
    previous_day: int | None = None
    for row_idx, col_idx, day_number in sorted(header_candidates):
        if previous_day is not None and day_number < previous_day:
            current_month += 1
        previous_day = day_number
        header_cells.append((row_idx, col_idx, date(YEAR, current_month, day_number).isoformat()))

    header_rows = sorted({row for row, _, _ in header_cells})
    next_header_by_row = {
        row: next((candidate for candidate in header_rows if candidate > row), ws.max_row + 1)
        for row in header_rows
    }

    for header_row, col_idx, event_date in header_cells:
        header_text = compact_text(ws.cell(header_row, col_idx).value)
        end_row = next_header_by_row[header_row]
        row_idx = header_row + 1
        while row_idx < end_row:
            label = compact_text(ws.cell(row_idx, col_idx).value).upper().rstrip(":")
            if label != "LOCATION":
                row_idx += 1
                continue
            location = compact_text(ws.cell(row_idx, col_idx + 1).value)
            time_label = compact_text(ws.cell(row_idx + 1, col_idx).value).upper().rstrip(":") if row_idx + 1 < end_row else ""
            time_value = compact_text(ws.cell(row_idx + 1, col_idx + 1).value) if row_idx + 1 < end_row else ""
            req_label = compact_text(ws.cell(row_idx + 2, col_idx).value).upper().rstrip(":") if row_idx + 2 < end_row else ""
            requirement = compact_text(ws.cell(row_idx + 2, col_idx + 1).value) if row_idx + 2 < end_row else ""
            if location or time_value or requirement:
                value_cells = [
                    style_ws.cell(row_idx, col_idx + 1),
                    style_ws.cell(row_idx + 1, col_idx + 1),
                    style_ws.cell(row_idx + 2, col_idx + 1),
                ]
                class_key, medical_day, medical_color = medical_owner_from_color(
                    event_date=event_date,
                    header_text=header_text,
                    value_cells=value_cells,
                    class_by_date_day=class_by_date_day,
                )
                if medical_day is not None and medical_day < 0:
                    row_idx += 3
                    continue
                title_parts = ["Medical coverage"]
                if location:
                    title_parts.append(location)
                if time_label != "TIME":
                    time_value = ""
                if req_label != "REQUIREMENT":
                    requirement = ""
                original_requirement = requirement
                requirement = normalize_ruck_medical_requirement(requirement, location, time_value)
                title = " - ".join(title_parts)
                if time_value:
                    title = f"{time_value} {title}"
                source_extra = {
                    "notesCell": f"{get_column_letter(col_idx + 1)}{row_idx + 2}",
                    "notesRow": row_idx + 2,
                    "notesCol": col_idx + 1,
                    "medicalHeader": header_text,
                    "medicalDay": medical_day,
                    "medicalFontColor": medical_color,
                }
                if original_requirement and original_requirement != requirement:
                    source_extra["originalNotes"] = original_requirement
                make_event(
                    events=events,
                    source_file=path.name,
                    sheet=ws.title,
                    row=row_idx,
                    col=col_idx,
                    event_date=event_date,
                    title=title,
                    header="Medical Coverage",
                    assigned=requirement,
                    source_kind="medical",
                    class_key=class_key,
                    location=location,
                    source_extra=source_extra,
                )
            row_idx += 3


def meal_name(code: str) -> str:
    upper = code.upper()
    return {"B": "Breakfast", "L": "Lunch", "D": "Dinner"}.get(upper, upper.title())


def parse_mess_matrix(path: Path, events: list[dict[str, Any]], notes: list[dict[str, Any]]) -> None:
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                text = compact_text(cell.value)
                if re.search(r"AIR\s*ASSAULT|AIRASSAULT|AASLT|AIR\s*ASLT", text, re.I):
                    notes.append(
                        {
                            "file": path.name,
                            "sheet": ws.title,
                            "cell": cell.coordinate,
                            "text": text,
                        }
                    )

        date_by_col: dict[int, str] = {}
        meal_by_col: dict[int, str] = {}
        for col_idx in range(1, ws.max_column + 1):
            for row_idx in range(1, min(ws.max_row, 5) + 1):
                date_value = date_to_iso(ws.cell(row_idx, col_idx).value, ws.title)
                if date_value:
                    for offset in range(0, 3):
                        if col_idx + offset <= ws.max_column:
                            date_by_col[col_idx + offset] = date_value
            meal_value = compact_text(ws.cell(3, col_idx).value).upper()
            if meal_value in {"B", "L", "D"}:
                meal_by_col[col_idx] = meal_name(meal_value)

        for row_idx in range(1, ws.max_row + 1):
            unit = compact_text(ws.cell(row_idx, 1).value)
            if not re.search(r"AIR\s*ASSAULT|AIRASSAULT|AASLT|AIR\s*ASLT", unit, re.I):
                continue
            if "TOTAL" in unit.upper():
                continue
            for col_idx in range(4, ws.max_column + 1):
                value = compact_text(ws.cell(row_idx, col_idx).value).upper()
                if not value:
                    continue
                if value not in {"CM", "M", "M+", "HM+", "AF"}:
                    continue
                event_date = date_by_col.get(col_idx)
                meal = meal_by_col.get(col_idx, "Meal")
                if not event_date:
                    continue
                if value == "CM":
                    meal_status = "Mess Hall"
                elif value in {"M", "M+", "HM+"}:
                    meal_status = "MRE"
                elif value == "AF":
                    meal_status = "A-Frame"
                else:
                    meal_status = value
                title = f"{unit} {meal}: {meal_status} ({value})"
                location = meal_status if meal_status in {"Mess Hall", "A-Frame"} else ""
                make_event(
                    events=events,
                    source_file=path.name,
                    sheet=ws.title,
                    row=row_idx,
                    col=col_idx,
                    event_date=event_date,
                    title=title,
                    header="Mess Matrix",
                    assigned="",
                    source_kind="mess",
                    class_key=infer_class_key(unit),
                    location=location,
                )


def dedupe_events(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[tuple[Any, ...], dict[str, Any]] = {}
    for event in events:
        key = dedupe_key(event)
        existing = merged.get(key)
        if existing:
            merge_event(existing, event)
            continue
        merged[key] = event
    cleaned = list(merged.values())
    cleaned.sort(key=lambda e: (e["date"], e.get("start") or "99:99", e["title"]))
    for idx, event in enumerate(cleaned, start=1):
        event["id"] = f"evt-{idx:05d}"
    return cleaned


def dedupe_taskings(taskings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[tuple[Any, ...], dict[str, Any]] = {}
    for tasking in taskings:
        key = (
            tasking.get("date"),
            tasking.get("start") or "",
            tasking.get("end") or "",
            tasking.get("track") or "",
            normalize_key_text(tasking.get("title", "")),
            normalize_key_text(tasking.get("location", "")),
        )
        existing = merged.get(key)
        if existing:
            sources = existing.setdefault("sources", [existing["source"]])
            if tasking["source"] not in sources:
                sources.append(tasking["source"])
            continue
        merged[key] = tasking
    cleaned = list(merged.values())
    cleaned.sort(key=lambda item: (item["date"], item.get("start") or "99:99", item.get("track") or "", item["title"]))
    return cleaned


def dedupe_support_items(support_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[tuple[Any, ...], dict[str, Any]] = {}
    for item in support_items:
        key = (
            item.get("lane"),
            item.get("item"),
            item.get("datetime"),
            item.get("track") or "",
        )
        existing = merged.get(key)
        if existing:
            sources = existing.setdefault("sources", [existing["source"]])
            if item["source"] not in sources:
                sources.append(item["source"])
            continue
        merged[key] = item
    cleaned = list(merged.values())
    cleaned.sort(key=lambda item: (item.get("datetime") or "", item.get("lane") or "", item.get("item") or ""))
    return cleaned


def normalize_key_text(text: str) -> str:
    upper = compact_text(text).upper()
    upper = re.sub(r"\b[0-2]?\d[0-5]\d\s*-\s*(?:[0-2]?\d[0-5]\d|UTC)\b", "", upper)
    return re.sub(r"\s+", " ", upper).strip()


def dedupe_title_key(event: dict[str, Any]) -> str:
    title = compact_text(event.get("title", ""))
    if event.get("category") != "Meals":
        return normalize_key_text(title)
    labels = meal_labels(title)
    if not labels:
        return normalize_key_text(title)
    status_text = f"{title} {event.get('notes', '')}".upper()
    status = ""
    if "MRE" in status_text:
        status = "MRE"
    elif "MESS HALL" in status_text or "(CM)" in status_text:
        status = "CM"
    elif "A-FRAME" in status_text or "(AF)" in status_text:
        status = "AF"
    return f"{'/'.join(labels)}|{status}"


def dedupe_key(event: dict[str, Any]) -> tuple[Any, ...]:
    return (
        event.get("date"),
        event.get("category"),
        event.get("classKey") or "",
        event.get("start") or "",
        event.get("end") or "",
        normalize_key_text(event.get("location", "")),
        dedupe_title_key(event),
    )


def merge_event(existing: dict[str, Any], event: dict[str, Any]) -> None:
    if event.get("start") and not existing.get("start"):
        existing["start"] = event["start"]
    if event.get("end") and not existing.get("end"):
        existing["end"] = event["end"]
    if event.get("location") and not existing.get("location"):
        existing["location"] = event["location"]
    if event.get("classKey") and not existing.get("classKey"):
        existing["classKey"] = event["classKey"]
    if event.get("notes") and event["notes"] not in existing.get("notes", ""):
        existing["notes"] = "; ".join(part for part in [existing.get("notes", ""), event["notes"]] if part)

    sources = existing.setdefault("sources", [existing["source"]])
    if event["source"] not in sources:
        sources.append(event["source"])


def normalize_medical_ruck_labels(events: list[dict[str, Any]]) -> None:
    ruck_by_date: dict[str, str] = {}
    ruck_by_date_class: dict[tuple[str, str], str] = {}
    for event in events:
        if event.get("category") == "Medical":
            continue
        text = f"{event.get('title', '')} {event.get('location', '')}".upper()
        label = ""
        if "12 MILE RUCK" in text:
            label = "12 Mile Ruck Start/Finish"
        elif "6 MILE RUCK" in text and event["date"] not in ruck_by_date:
            label = "6 Mile Ruck Start/Finish"
        if not label:
            continue
        class_key = event.get("classKey") or ""
        if class_key:
            current = ruck_by_date_class.get((event["date"], class_key))
            if not current or label.startswith("12 Mile"):
                ruck_by_date_class[(event["date"], class_key)] = label
        if label.startswith("12 Mile") or event["date"] not in ruck_by_date:
            ruck_by_date[event["date"]] = label

    for event in events:
        if event.get("category") != "Medical":
            continue
        text = f"{event.get('title', '')} {event.get('location', '')}".upper()
        if "RUCK S/F" not in text:
            continue
        label = ruck_by_date_class.get((event["date"], event.get("classKey") or "")) or ruck_by_date.get(event["date"], "Ruck Start/Finish")
        event["title"] = re.sub(r"RUCK\s+S/F", label, event["title"], flags=re.I)
        event["location"] = label


def summarize(events: list[dict[str, Any]]) -> dict[str, Any]:
    by_date: defaultdict[str, int] = defaultdict(int)
    by_category: defaultdict[str, int] = defaultdict(int)
    for event in events:
        by_date[event["date"]] += 1
        by_category[event["category"]] += 1
    return {
        "dateRange": [min(by_date) if by_date else None, max(by_date) if by_date else None],
        "eventCount": len(events),
        "sourceCount": len(SOURCE_FILES),
        "peopleCount": 0,
        "eventsByDate": dict(sorted(by_date.items())),
        "eventsByCategory": dict(sorted(by_category.items())),
        "people": [],
    }


def stable_id(prefix: str, *parts: Any) -> str:
    digest = sha1("|".join(compact_text(part) for part in parts).encode("utf-8")).hexdigest()[:12]
    return f"{prefix}_{digest}"


def source_refs_for_event(
    event: dict[str, Any],
    source_by_name: dict[str, str],
    sheet_by_key: dict[tuple[str, str], str],
) -> list[dict[str, Any]]:
    source_items = event.get("sources") or [event.get("source")]
    refs = []
    seen = set()
    for source in source_items:
        if not source:
            continue
        file_name = source.get("file", "")
        sheet_name = source.get("sheet", "")
        address = source.get("cell", "")
        key = (file_name, sheet_name, address, source.get("row"), source.get("col"))
        if key in seen:
            continue
        seen.add(key)
        source_id = source_by_name.get(file_name, stable_id("src", file_name))
        sheet_id = sheet_by_key.get((file_name, sheet_name), stable_id("sheet", file_name, sheet_name))
        refs.append(
            {
                "sourceId": source_id,
                "sheetId": sheet_id,
                "cellId": stable_id("cell", file_name, sheet_name, address),
                "file": file_name,
                "sheet": sheet_name,
                "address": address,
                "row": source.get("row"),
                "col": source.get("col"),
                "notesCell": source.get("notesCell"),
                "originalNotes": source.get("originalNotes"),
                "medicalHeader": source.get("medicalHeader"),
                "medicalDay": source.get("medicalDay"),
                "medicalFontColor": source.get("medicalFontColor"),
            }
        )
    return refs


def issue(
    *,
    diagnostics: list[dict[str, Any]],
    severity: str,
    kind: str,
    message: str,
    date_value: str | None = None,
    class_key: str | None = None,
    event_ids: list[str] | None = None,
    source_refs: list[dict[str, Any]] | None = None,
) -> None:
    diagnostics.append(
        {
            "id": stable_id("issue", kind, message, date_value or "", class_key or "", ",".join(event_ids or [])),
            "severity": severity,
            "kind": kind,
            "date": date_value,
            "classKey": class_key,
            "message": message,
            "eventIds": event_ids or [],
            "sourceRefs": source_refs or [],
        }
    )


def meal_label_for_event(event: dict[str, Any]) -> str:
    labels = meal_labels(f"{event.get('title', '')} {event.get('notes', '')}")
    return labels[0] if labels else ""


def event_counts_as_active_track(event: dict[str, Any]) -> bool:
    if not event.get("classKey"):
        return False
    role = compact_text(event.get("role", "")).upper()
    if role in {"ADMIN", "CADRE"}:
        return False
    if event.get("category") in {"Cadre", "Logistics", "Operations"}:
        return False
    return True


def build_diagnostics(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    diagnostics: list[dict[str, Any]] = []
    tracks_by_date: defaultdict[str, set[str]] = defaultdict(set)
    scheduled_meals_by_key: defaultdict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    medical_by_window: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    scheduled_meal_keys = set()
    mess_meal_keys = set()

    for event in events:
        date_value = event.get("date")
        class_key = event.get("classKey") or ""
        if date_value and class_key and event_counts_as_active_track(event):
            tracks_by_date[date_value].add(class_key)

        if event.get("category") == "Meals":
            meal = meal_label_for_event(event)
            if meal and class_key:
                key = (date_value, class_key, meal)
                if event.get("sourceKind") == "mess":
                    mess_meal_keys.add(key)
                else:
                    scheduled_meal_keys.add(key)
                    scheduled_meals_by_key[key].append(event)

        if event.get("category") == "Medical" and date_value and event.get("start") and not class_key:
            medical_by_window[(date_value, event.get("start"))].append(event)

        if event.get("sourceKind") not in {"medical", "mess"} and event.get("category") in {"Training", "Instruction", "Meals"} and not class_key:
            issue(
                diagnostics=diagnostics,
                severity="medium",
                kind="missing-track",
                date_value=date_value,
                message=f"No Air Assault track was inferred for {event.get('title')}.",
                event_ids=[event.get("id", "")],
                source_refs=event.get("sourceRefs", []),
            )

        if event.get("category") == "Medical" and not event.get("start"):
            issue(
                diagnostics=diagnostics,
                severity="low",
                kind="missing-time",
                date_value=date_value,
                class_key=class_key or None,
                message=f"No start time was found for {event.get('title')}.",
                event_ids=[event.get("id", "")],
                source_refs=event.get("sourceRefs", []),
            )

        original_notes = next((ref.get("originalNotes") for ref in event.get("sourceRefs", []) if ref.get("originalNotes")), "")
        if original_notes:
            issue(
                diagnostics=diagnostics,
                severity="info",
                kind="normalized-medical-requirement",
                date_value=date_value,
                message=f"Ruck FLA requirement normalized from '{original_notes}' to '{event.get('notes')}'.",
                event_ids=[event.get("id", "")],
                source_refs=event.get("sourceRefs", []),
            )

    for date_value, tracks in tracks_by_date.items():
        if len(tracks) > 2:
            issue(
                diagnostics=diagnostics,
                severity="high",
                kind="too-many-active-tracks",
                date_value=date_value,
                message=f"More than two Air Assault tracks appear active: {', '.join(sorted(tracks))}.",
            )

    for (date_value, class_key, meal), meal_events in scheduled_meals_by_key.items():
        if len(meal_events) > 1:
            times = ", ".join(sorted({event.get("start") or "no time" for event in meal_events}))
            issue(
                diagnostics=diagnostics,
                severity="medium",
                kind="duplicate-meal-rows",
                date_value=date_value,
                class_key=class_key,
                message=f"{class_key} has {len(meal_events)} {meal} rows before display merge ({times}).",
                event_ids=[event.get("id", "") for event in meal_events],
                source_refs=[ref for event in meal_events for ref in event.get("sourceRefs", [])],
            )

    for (date_value, start), medical_events in medical_by_window.items():
        locations = sorted({event.get("location") or event.get("title") for event in medical_events})
        if len(locations) > 1:
            issue(
                diagnostics=diagnostics,
                severity="info",
                kind="shared-medical-same-time",
                date_value=date_value,
                message=f"Shared medical coverage has multiple locations at {start}: {', '.join(locations)}.",
                event_ids=[event.get("id", "") for event in medical_events],
                source_refs=[ref for event in medical_events for ref in event.get("sourceRefs", [])],
            )

    mess_only = sorted(mess_meal_keys - scheduled_meal_keys)
    if mess_only:
        issue(
            diagnostics=diagnostics,
            severity="low",
            kind="mess-only-meal-summary",
            message=f"{len(mess_only)} mess matrix meal statuses have no matching schedule time/location row. They remain usable as meal-type authority, with default times only when needed.",
        )

    diagnostics.sort(key=lambda item: (item.get("date") or "9999", item["severity"], item["kind"], item["message"]))
    return diagnostics


def build_database(payload: dict[str, Any]) -> dict[str, Any]:
    source_records: list[dict[str, Any]] = []
    sheet_records: list[dict[str, Any]] = []
    cell_records: list[dict[str, Any]] = []
    source_by_name: dict[str, str] = {}
    sheet_by_key: dict[tuple[str, str], str] = {}

    for source_index, source in enumerate(payload["sources"], start=1):
        source_id = stable_id("src", source["name"])
        source_by_name[source["name"]] = source_id
        source_records.append({"id": source_id, "name": source["name"], "workbookIndex": source_index})
        for sheet_index, sheet in enumerate(source["sheets"], start=1):
            sheet_id = stable_id("sheet", source["name"], sheet["name"])
            sheet_by_key[(source["name"], sheet["name"])] = sheet_id
            sheet_records.append(
                {
                    "id": sheet_id,
                    "sourceId": source_id,
                    "name": sheet["name"],
                    "sheetIndex": sheet_index,
                    "maxRow": sheet["maxRow"],
                    "maxColumn": sheet["maxColumn"],
                }
            )
            for cell in sheet["cells"]:
                cell_records.append(
                    {
                        "id": stable_id("cell", source["name"], sheet["name"], cell["a"]),
                        "sourceId": source_id,
                        "sheetId": sheet_id,
                        "address": cell["a"],
                        "row": cell["r"],
                        "col": cell["c"],
                        "value": cell.get("v"),
                        "formula": cell.get("f"),
                    }
                )

    event_records = []
    for event in payload["events"]:
        source_refs = source_refs_for_event(event, source_by_name, sheet_by_key)
        primary_ref = source_refs[0] if source_refs else None
        clean_event = {
            key: value
            for key, value in event.items()
            if key not in {"source", "sources", "relatedSources"}
        }
        clean_event["sourceRefs"] = source_refs
        clean_event["source"] = primary_ref
        event_records.append(clean_event)

    tasking_records = []
    for tasking in payload.get("taskings", []):
        source_refs = source_refs_for_event(tasking, source_by_name, sheet_by_key)
        primary_ref = source_refs[0] if source_refs else None
        clean_tasking = {
            key: value
            for key, value in tasking.items()
            if key not in {"source", "sources", "relatedSources"}
        }
        clean_tasking["sourceRefs"] = source_refs
        clean_tasking["source"] = primary_ref
        tasking_records.append(clean_tasking)

    support_records = []
    for item in payload.get("supportItems", []):
        source_refs = source_refs_for_event(item, source_by_name, sheet_by_key)
        primary_ref = source_refs[0] if source_refs else None
        clean_item = {
            key: value
            for key, value in item.items()
            if key not in {"source", "sources", "relatedSources"}
        }
        clean_item["sourceRefs"] = source_refs
        clean_item["source"] = primary_ref
        support_records.append(clean_item)

    note_records = []
    for note in payload["notes"]:
        refs = source_refs_for_event({"source": note}, source_by_name, sheet_by_key)
        note_records.append({**note, "id": stable_id("note", note.get("file"), note.get("sheet"), note.get("cell"), note.get("text")), "sourceRefs": refs})

    diagnostics = build_diagnostics(event_records)
    severity_counts: defaultdict[str, int] = defaultdict(int)
    for item in diagnostics:
        severity_counts[item["severity"]] += 1

    summary = {
        **payload["summary"],
        "database": {
            "sourceCount": len(source_records),
            "sheetCount": len(sheet_records),
            "cellCount": len(cell_records),
            "diagnosticCount": len(diagnostics),
            "diagnosticsBySeverity": dict(sorted(severity_counts.items())),
        },
    }

    return {
        "schemaVersion": 1,
        "generatedAt": payload["generatedAt"],
        "operation": payload["operation"],
        "summary": summary,
        "tables": {
            "events": event_records,
            "sources": source_records,
            "sheets": sheet_records,
            "cells": cell_records,
            "notes": note_records,
            "taskings": tasking_records,
            "supportItems": support_records,
            "diagnostics": diagnostics,
        },
    }


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    raw_sources = []
    events: list[dict[str, Any]] = []
    taskings: list[dict[str, Any]] = []
    support_items: list[dict[str, Any]] = []
    notes: list[dict[str, Any]] = []
    for path in SOURCE_FILES:
        raw_sources.append(extract_raw_workbook(path))
        if path == LRTC_SOURCE_FILE or path.name.startswith("Total AASLT Cadre LRTC"):
            parse_lrtc(path, events, taskings, support_items)
        elif path.name == "1. CST 26 Mess Matrix (Live Document).xlsx":
            parse_mess_matrix(path, events, notes)
    normalize_medical_ruck_labels(events)
    events = dedupe_events(events)
    taskings = dedupe_taskings(taskings)
    support_items = dedupe_support_items(support_items)
    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "operation": {
            "name": "West Point Air Assault",
            "dashboardTitle": "AASLT Control Dashboard",
            "year": YEAR,
        },
        "summary": summarize(events),
        "events": events,
        "taskings": taskings,
        "supportItems": support_items,
        "notes": notes,
        "sources": raw_sources,
    }
    database = build_database(payload)
    db_path = DATA_DIR / "aaslt-db.json"
    legacy_path = DATA_DIR / "schedule.json"
    db_path.write_text(json.dumps(database, indent=2), encoding="utf-8")
    legacy_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {db_path}")
    print(f"Wrote legacy export {legacy_path}")
    print(json.dumps(database["summary"], indent=2))


if __name__ == "__main__":
    main()
